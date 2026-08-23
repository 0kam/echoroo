#!/usr/bin/env python3
"""Build the bundled vernacular-name data files shipped in ``echoroo.data``.

This is a **developer tool**, not part of the runtime. It reads three upstream
inputs and emits the deterministic bundle consumed by
``echoroo.services.vernacular_bundle``:

``ioc_ja.csv`` / ``ioc_ja.meta.json``
    Japanese species names extracted from the IOC World Bird List
    (Multilingual Version). Primary display source for ``locale="ja"``.

``birdnet_crosswalk.csv`` / ``birdnet_crosswalk.meta.json``
    BirdNET V2.4 label scientific name → AviList scientific name, plus the
    AvibaseID and Cornell species code. BirdNET labels follow eBird/Clements
    while the IOC list follows its own taxonomy, so recent genus revisions
    (``Accipiter`` → ``Tachyspiza`` / ``Astur``, ``Charadrius`` →
    ``Anarhynchus``, ...) would otherwise silently drop a Japanese name.

``birdnet_unresolved.txt``
    BirdNET labels no rule matched (non-birds such as ``Engine`` / ``Dog`` and
    a handful of genuinely unmatched taxa), kept for manual curation.

``overrides.csv``
    Hand-maintained ``birdnet_scientific_name,avilist_scientific_name`` pairs.
    Read by this script when present; never rewritten by it.

Matching order per BirdNET label:

1. exact ``scientific_name`` match against AviList species rows;
2. BirdNET common name == ``English_name_Clements_v2025``, but only when that
   English name is *unique* among AviList species rows (ambiguous names are
   left unresolved rather than guessed);
3. an explicit ``overrides.csv`` entry.

Determinism: identical inputs produce byte-identical outputs. All rows are
sorted, CSVs are written with ``\\n`` line endings and no BOM, and the JSON
sidecars use sorted keys.

Usage (``openpyxl`` is intentionally **not** a project dependency — it is only
needed to build the bundle, not to serve it)::

    uv run --with openpyxl python scripts/build_vernacular_bundle.py \\
        --ioc /path/to/Multiling\\ IOC\\ 15.2.xlsx \\
        --avilist /path/to/AviList-v2025b-extended.xlsx \\
        --birdnet-labels /path/to/BirdNET_GLOBAL_6K_V2.4_Labels.txt \\
        --out-dir echoroo/data/vernacular

Or, with openpyxl available on an external path::

    PYTHONPATH=/path/to/pylib python3 scripts/build_vernacular_bundle.py ...
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Upstream dataset provenance (mirrored into the *.meta.json sidecars)
# ---------------------------------------------------------------------------

IOC_DATASET = "IOC World Bird List v15.2 — Multilingual Version"
IOC_VERSION = "15.2"
IOC_LICENSE = "CC BY 3.0 Unported"
IOC_CITATION = (
    "Gill F, D Donsker & P Rasmussen (Eds). 2026. IOC World Bird List (v15.2). "
    "doi:10.14344/IOC.ML.15.2"
)
IOC_URL = "https://worldbirdnames.org/Multiling%20IOC%2015.2.xlsx"
IOC_DOI = "10.14344/IOC.ML.15.2"

AVILIST_DATASET = "AviList: The Global Avian Checklist, v2025b (extended)"
AVILIST_VERSION = "v2025b"
AVILIST_LICENSE = "CC BY 4.0"
AVILIST_CITATION = (
    "AviList Core Team. 2026. AviList: The Global Avian Checklist, v2025b. "
    "https://doi.org/10.2173/avilist.v2025b"
)
AVILIST_URL = (
    "https://www.avilist.org/wp-content/uploads/2026/06/"
    "AviList-v2025b-10Jun2026-extended.xlsx"
)
AVILIST_DOI = "10.2173/avilist.v2025b"

BIRDNET_LABEL_SET_VERSION = "2.4"

# Date the upstream files were downloaded. Bump alongside the inputs.
RETRIEVED = "2026-08-21"

# Output file names (relative to --out-dir).
IOC_CSV_NAME = "ioc_ja.csv"
IOC_META_NAME = "ioc_ja.meta.json"
CROSSWALK_CSV_NAME = "birdnet_crosswalk.csv"
CROSSWALK_META_NAME = "birdnet_crosswalk.meta.json"
UNRESOLVED_NAME = "birdnet_unresolved.txt"
OVERRIDES_NAME = "overrides.csv"

# Column headers of the IOC "List" sheet we care about.
_IOC_SHEET = "List"
_IOC_SCIENTIFIC_HEADER_PREFIX = "IOC_"
_IOC_JAPANESE_HEADER = "Japanese"

# Match-method labels emitted into birdnet_crosswalk.csv.
MATCH_EXACT = "exact"
MATCH_CLEMENTS_ENGLISH = "clements_english"
MATCH_OVERRIDE = "override"


# ---------------------------------------------------------------------------
# Pure data structures
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class AviListSpecies:
    """One AviList row with ``Taxon_rank == "species"``."""

    scientific_name: str
    english_name_clements: str = ""
    cornell_code: str = ""
    avibase_id: str = ""


@dataclass(frozen=True)
class CrosswalkRow:
    """A resolved BirdNET label → AviList species mapping."""

    birdnet_scientific_name: str
    avilist_scientific_name: str
    avibase_id: str
    cornell_code: str
    match_method: str

    def as_row(self) -> list[str]:
        return [
            self.birdnet_scientific_name,
            self.avilist_scientific_name,
            self.avibase_id,
            self.cornell_code,
            self.match_method,
        ]


@dataclass
class MatchResult:
    """Outcome of matching the BirdNET label set against AviList."""

    rows: list[CrosswalkRow] = field(default_factory=list)
    unresolved: list[tuple[str, str]] = field(default_factory=list)

    @property
    def method_counts(self) -> dict[str, int]:
        counts = {MATCH_EXACT: 0, MATCH_CLEMENTS_ENGLISH: 0, MATCH_OVERRIDE: 0}
        for row in self.rows:
            counts[row.match_method] = counts.get(row.match_method, 0) + 1
        return counts


# ---------------------------------------------------------------------------
# Pure matching logic (unit-tested without any xlsx input)
# ---------------------------------------------------------------------------


def parse_birdnet_label(label: str) -> tuple[str, str]:
    """Parse a BirdNET label into ``(scientific_name, common_name)``.

    Mirrors ``echoroo.services.taxon_seeder._parse_birdnet_label`` exactly so
    the crosswalk keys line up with the ``taxa.scientific_name`` values the
    seeder writes. Both shipped label formats are handled: the packaged
    ``Genus_species_Common Name`` form and the model-directory
    ``Genus species_Common Name`` form both yield ``"Genus species"``.
    """
    parts = label.split("_", 2)
    if len(parts) == 3:
        scientific_name = f"{parts[0]} {parts[1]}"
        common_name = parts[2]
    elif len(parts) == 2:
        scientific_name = parts[0]
        common_name = parts[1]
    else:
        scientific_name = label
        common_name = ""
    return scientific_name.strip(), common_name.strip()


def parse_birdnet_labels(lines: Iterable[str]) -> list[tuple[str, str]]:
    """Parse a BirdNET label file body into ``(scientific, common)`` pairs.

    Blank lines are dropped; labels without a scientific-name component are
    skipped. Input order is preserved (output ordering is applied later).
    """
    parsed: list[tuple[str, str]] = []
    for raw in lines:
        label = raw.strip()
        if not label:
            continue
        scientific_name, common_name = parse_birdnet_label(label)
        if not scientific_name:
            continue
        parsed.append((scientific_name, common_name))
    return parsed


def build_clements_english_index(
    species: Iterable[AviListSpecies],
) -> dict[str, AviListSpecies]:
    """Index AviList species by Clements English name, dropping ambiguities.

    An English name shared by two or more AviList species cannot identify a
    single concept, so it is removed from the index entirely: rule 2 must never
    guess. Comparison is case-insensitive and whitespace-normalised.
    """
    index: dict[str, AviListSpecies] = {}
    ambiguous: set[str] = set()
    for entry in species:
        key = _normalize_english(entry.english_name_clements)
        if not key:
            continue
        if key in index:
            ambiguous.add(key)
            continue
        index[key] = entry
    for key in ambiguous:
        index.pop(key, None)
    return index


def _normalize_english(name: str) -> str:
    return " ".join(name.split()).casefold()


def match_birdnet_labels(
    labels: Sequence[tuple[str, str]],
    species: Sequence[AviListSpecies],
    overrides: Mapping[str, str] | None = None,
) -> MatchResult:
    """Resolve BirdNET labels onto AviList species concepts.

    Args:
        labels: ``(scientific_name, common_name)`` pairs from the BirdNET
            label set.
        species: AviList rows with ``Taxon_rank == "species"``.
        overrides: Optional ``birdnet_scientific_name ->
            avilist_scientific_name`` fallback map. Entries pointing at a
            scientific name absent from ``species`` are ignored.

    Returns:
        A :class:`MatchResult` whose ``rows`` are sorted by
        ``(birdnet_scientific_name, avilist_scientific_name)`` and whose
        ``unresolved`` entries are sorted by ``(scientific, common)``. Rows are
        emitted even when both scientific names are identical, because the
        AvibaseID / Cornell code carried alongside is useful downstream.
    """
    by_scientific = {entry.scientific_name: entry for entry in species}
    by_english = build_clements_english_index(species)
    override_map = dict(overrides or {})

    rows: list[CrosswalkRow] = []
    unresolved: list[tuple[str, str]] = []
    seen: set[str] = set()

    for scientific_name, common_name in labels:
        if scientific_name in seen:
            # Duplicate labels would produce duplicate crosswalk keys; the
            # first occurrence wins so the output stays a strict 1:1 map.
            continue
        seen.add(scientific_name)

        match: AviListSpecies | None = None
        method = ""

        exact = by_scientific.get(scientific_name)
        if exact is not None:
            match, method = exact, MATCH_EXACT
        else:
            by_common = by_english.get(_normalize_english(common_name))
            if by_common is not None:
                match, method = by_common, MATCH_CLEMENTS_ENGLISH
            else:
                override_target = override_map.get(scientific_name)
                if override_target is not None:
                    overridden = by_scientific.get(override_target)
                    if overridden is not None:
                        match, method = overridden, MATCH_OVERRIDE

        if match is None:
            unresolved.append((scientific_name, common_name))
            continue

        rows.append(
            CrosswalkRow(
                birdnet_scientific_name=scientific_name,
                avilist_scientific_name=match.scientific_name,
                avibase_id=match.avibase_id,
                cornell_code=match.cornell_code,
                match_method=method,
            )
        )

    rows.sort(key=lambda row: (row.birdnet_scientific_name, row.avilist_scientific_name))
    unresolved.sort()
    return MatchResult(rows=rows, unresolved=unresolved)


def sort_name_rows(rows: Iterable[tuple[str, str]]) -> list[tuple[str, str]]:
    """Deduplicate and sort ``(scientific_name, name)`` pairs.

    The first occurrence of a scientific name wins so an upstream duplicate row
    cannot make the output depend on sort stability.
    """
    seen: dict[str, str] = {}
    for scientific_name, name in rows:
        key = scientific_name.strip()
        value = name.strip()
        if not key or not value or key in seen:
            continue
        seen[key] = value
    return sorted(seen.items())


# ---------------------------------------------------------------------------
# Input readers (xlsx / txt / csv)
# ---------------------------------------------------------------------------


def _load_openpyxl() -> Any:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - environment dependent
        raise SystemExit(
            "openpyxl is required to build the vernacular bundle but is not "
            "installed. It is intentionally not a project dependency (build-"
            "time only). Run this script via "
            "`uv run --with openpyxl python scripts/build_vernacular_bundle.py "
            "...` or point PYTHONPATH at an environment that provides it."
        ) from exc
    return openpyxl


def read_ioc_japanese_names(path: Path) -> list[tuple[str, str]]:
    """Extract ``(scientific_name, japanese_name)`` from the IOC workbook.

    The ``List`` sheet's scientific-name column is version-stamped
    (``IOC_15.2``), so it is located by prefix rather than by exact header.
    Rows without a Japanese name are skipped.
    """
    openpyxl = _load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[_IOC_SHEET]
        rows = sheet.iter_rows(values_only=True)
        header = [_cell_text(cell) for cell in next(rows)]
        scientific_idx = _header_index_by_prefix(header, _IOC_SCIENTIFIC_HEADER_PREFIX)
        japanese_idx = _header_index(header, _IOC_JAPANESE_HEADER)

        extracted: list[tuple[str, str]] = []
        for row in rows:
            scientific_name = _cell_text(_at(row, scientific_idx))
            japanese_name = _cell_text(_at(row, japanese_idx))
            if not scientific_name or not japanese_name:
                continue
            extracted.append((scientific_name, japanese_name))
        return extracted
    finally:
        workbook.close()


def read_avilist_species(path: Path) -> list[AviListSpecies]:
    """Extract the ``Taxon_rank == "species"`` rows from the AviList workbook.

    The extended workbook carries a machine header row followed by a
    human-readable header row; the latter is skipped naturally because its
    ``Taxon_rank`` cell reads ``RANK`` rather than ``species``.
    """
    openpyxl = _load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        header = [_cell_text(cell) for cell in next(rows)]
        rank_idx = _header_index(header, "Taxon_rank")
        scientific_idx = _header_index(header, "Scientific_name")
        clements_idx = _header_index(header, "English_name_Clements_v2025")
        code_idx = _header_index(header, "Species_code_Cornell_Lab")
        avibase_idx = _header_index(header, "AvibaseID")

        species: list[AviListSpecies] = []
        for row in rows:
            if _cell_text(_at(row, rank_idx)).casefold() != "species":
                continue
            scientific_name = _cell_text(_at(row, scientific_idx))
            if not scientific_name:
                continue
            species.append(
                AviListSpecies(
                    scientific_name=scientific_name,
                    english_name_clements=_cell_text(_at(row, clements_idx)),
                    cornell_code=_cell_text(_at(row, code_idx)),
                    avibase_id=_cell_text(_at(row, avibase_idx)),
                )
            )
        return species
    finally:
        workbook.close()


def read_overrides(path: Path) -> dict[str, str]:
    """Read ``overrides.csv`` if it exists; return an empty map otherwise."""
    if not path.exists():
        return {}
    overrides: dict[str, str] = {}
    with path.open(encoding="utf-8", newline="") as handle:
        for record in csv.DictReader(handle):
            source = (record.get("birdnet_scientific_name") or "").strip()
            target = (record.get("avilist_scientific_name") or "").strip()
            if source and target:
                overrides[source] = target
    return overrides


def _header_index(header: Sequence[str], name: str) -> int:
    try:
        return header.index(name)
    except ValueError as exc:
        raise SystemExit(f"Expected column {name!r} not found in {header!r}") from exc


def _header_index_by_prefix(header: Sequence[str], prefix: str) -> int:
    for index, value in enumerate(header):
        if value.startswith(prefix):
            return index
    raise SystemExit(f"No column starting with {prefix!r} found in {header!r}")


def _at(row: Sequence[object], index: int) -> object:
    return row[index] if index < len(row) else None


def _cell_text(cell: object) -> str:
    if cell is None:
        return ""
    return str(cell).strip()


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------


def write_csv(path: Path, header: Sequence[str], rows: Iterable[Sequence[str]]) -> int:
    """Write a UTF-8 (no BOM), LF-terminated CSV. Returns the row count."""
    count = 0
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerow(header)
        for row in rows:
            writer.writerow(row)
            count += 1
    return count


def write_json(path: Path, payload: Mapping[str, object]) -> None:
    """Write a deterministic JSON sidecar (sorted keys, trailing newline)."""
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def write_lines(path: Path, lines: Iterable[str]) -> None:
    body = "".join(f"{line}\n" for line in lines)
    path.write_text(body, encoding="utf-8")


def ensure_overrides_file(path: Path) -> None:
    """Create ``overrides.csv`` with a header and no rows when missing."""
    if path.exists():
        return
    write_csv(path, ["birdnet_scientific_name", "avilist_scientific_name"], [])


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def _parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build the bundled vernacular-name data files.",
    )
    parser.add_argument("--ioc", required=True, type=Path, help="IOC Multilingual xlsx")
    parser.add_argument(
        "--avilist", required=True, type=Path, help="AviList extended xlsx"
    )
    parser.add_argument(
        "--birdnet-labels",
        required=True,
        type=Path,
        help="BirdNET V2.4 English label file",
    )
    parser.add_argument(
        "--out-dir",
        required=True,
        type=Path,
        help="Destination directory (echoroo/data/vernacular)",
    )
    parser.add_argument(
        "--overrides",
        type=Path,
        default=None,
        help="Override CSV (defaults to <out-dir>/overrides.csv)",
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = _parse_args(argv)
    out_dir: Path = args.out_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    overrides_path: Path = args.overrides or (out_dir / OVERRIDES_NAME)
    ensure_overrides_file(overrides_path)
    overrides = read_overrides(overrides_path)

    # --- IOC Japanese names --------------------------------------------------
    ioc_rows = sort_name_rows(read_ioc_japanese_names(args.ioc))
    ioc_count = write_csv(
        out_dir / IOC_CSV_NAME, ["scientific_name", "name"], ioc_rows
    )
    write_json(
        out_dir / IOC_META_NAME,
        {
            "source": "ioc",
            "dataset": IOC_DATASET,
            "version": IOC_VERSION,
            "locale": "ja",
            "license": IOC_LICENSE,
            "citation": IOC_CITATION,
            "url": IOC_URL,
            "doi": IOC_DOI,
            "retrieved": RETRIEVED,
            "rows": ioc_count,
        },
    )

    # --- BirdNET → AviList crosswalk ----------------------------------------
    species = read_avilist_species(args.avilist)
    labels = parse_birdnet_labels(
        args.birdnet_labels.read_text(encoding="utf-8").splitlines()
    )
    result = match_birdnet_labels(labels, species, overrides)

    crosswalk_count = write_csv(
        out_dir / CROSSWALK_CSV_NAME,
        [
            "birdnet_scientific_name",
            "avilist_scientific_name",
            "avibase_id",
            "cornell_code",
            "match_method",
        ],
        (row.as_row() for row in result.rows),
    )
    write_lines(
        out_dir / UNRESOLVED_NAME,
        (f"{scientific}\t{common}" for scientific, common in result.unresolved),
    )
    counts = result.method_counts
    write_json(
        out_dir / CROSSWALK_META_NAME,
        {
            "source": "avilist",
            "dataset": AVILIST_DATASET,
            "version": AVILIST_VERSION,
            "license": AVILIST_LICENSE,
            "citation": AVILIST_CITATION,
            "url": AVILIST_URL,
            "doi": AVILIST_DOI,
            "retrieved": RETRIEVED,
            "birdnet_label_set_version": BIRDNET_LABEL_SET_VERSION,
            "birdnet_labels": len(labels),
            "resolved": crosswalk_count,
            "unresolved": len(result.unresolved),
            "match_methods": counts,
        },
    )

    print(f"IOC Japanese names: {ioc_count} rows -> {out_dir / IOC_CSV_NAME}")
    print(f"AviList species rows: {len(species)}")
    print(f"BirdNET labels: {len(labels)}")
    print(
        "Crosswalk resolved: "
        f"{crosswalk_count} "
        f"(exact={counts[MATCH_EXACT]}, "
        f"clements_english={counts[MATCH_CLEMENTS_ENGLISH]}, "
        f"override={counts[MATCH_OVERRIDE]})"
    )
    print(f"Crosswalk unresolved: {len(result.unresolved)}")
    return 0


if __name__ == "__main__":  # pragma: no cover - CLI entrypoint
    sys.exit(main())
