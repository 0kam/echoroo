#!/usr/bin/env python3
"""Convert the Check-List of Japanese Birds (8th ed.) XLSX into a loader CSV.

The Ornithological Society of Japan publishes the species list of the
日本鳥類目録改訂第8版 as an XLSX (sheet ``リスト``) whose terms allow
derived use but state no redistribution license. It is therefore NOT bundled
with Echoroo; an operator downloads it themselves and converts it with this
host-side tool, then loads the CSV inside the container with
``python -m echoroo.scripts.load_authority_checklist``.

    uv run --with openpyxl python scripts/convert_osj_checklist.py \\
        jpbirdlist8ed_ver1.xlsx --out osj8_ja.csv

Only rows whose ``カテゴリ`` is ``種`` (species) are exported — Part A
(natural distribution) and Part B (introduced) alike. Subspecies, genus,
family and order rows are skipped. Output columns: ``scientific_name,name``
(UTF-8, sorted, deterministic).
"""

from __future__ import annotations

import argparse
import csv
import sys
from collections.abc import Iterable, Sequence
from pathlib import Path

SHEET_NAME = "リスト"
HEADER = ("掲載順", "Part", "カテゴリ", "種番号", "亜種番号", "学名", "著者", "和名")
SPECIES_CATEGORY = "種"


def extract_species_rows(rows: Iterable[Sequence[object]]) -> list[tuple[str, str]]:
    """Pick ``(scientific_name, japanese_name)`` for species rows.

    ``rows`` are the raw sheet rows *including* the header row. The header is
    validated so a re-ordered upstream file fails loudly instead of silently
    mapping the wrong columns. Blank names are skipped; whitespace is
    stripped; output is sorted and de-duplicated by scientific name (a later
    duplicate wins, mirroring the list order).
    """
    it = iter(rows)
    try:
        header = tuple(str(c).strip() if c is not None else "" for c in next(it))
    except StopIteration as exc:
        raise ValueError("checklist sheet is empty") from exc
    if header[: len(HEADER)] != HEADER:
        raise ValueError(f"unexpected header {header[: len(HEADER)]!r}; expected {HEADER!r}")

    col_category = HEADER.index("カテゴリ")
    col_sci = HEADER.index("学名")
    col_ja = HEADER.index("和名")

    picked: dict[str, str] = {}
    for row in it:
        if len(row) <= col_ja:
            continue
        if str(row[col_category] or "").strip() != SPECIES_CATEGORY:
            continue
        sci = str(row[col_sci] or "").strip()
        ja = str(row[col_ja] or "").strip()
        if sci and ja:
            picked[sci] = ja
    return sorted(picked.items())


def write_csv(rows: Iterable[tuple[str, str]], out: Path) -> int:
    """Write ``scientific_name,name`` rows; return the row count."""
    count = 0
    with out.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerow(["scientific_name", "name"])
        for sci, ja in rows:
            writer.writerow([sci, ja])
            count += 1
    return count


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    parser.add_argument("xlsx", type=Path, help="Official OSJ checklist XLSX (sheet リスト)")
    parser.add_argument("--out", type=Path, required=True, help="Output CSV path")
    args = parser.parse_args(argv)

    try:
        import openpyxl
    except ImportError:
        print(
            "openpyxl is required: run with `uv run --with openpyxl python ...`",
            file=sys.stderr,
        )
        return 2

    workbook = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        print(f"sheet {SHEET_NAME!r} not found in {args.xlsx}", file=sys.stderr)
        return 2
    rows = extract_species_rows(workbook[SHEET_NAME].iter_rows(values_only=True))
    count = write_csv(rows, args.out)
    print(f"{count} species rows -> {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
